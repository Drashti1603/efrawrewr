import os
import openpyxl
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, jsonify
import tempfile

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('daily_expense.html')

@app.route('/daily-expense', methods=['GET', 'POST'])
def daily_expense():
    if request.method == 'POST':
        amount = request.form['amount']
        description = request.form['description']
        date = request.form['date']
        response = submit_expense(amount, description, date)
        return jsonify(response=response)
    return render_template('daily_expense.html')

@app.route('/document-scan', methods=['GET', 'POST'])
def document_scan():
    if request.method == 'POST':
        pan_number = request.form['pan_number']
        file = request.files['document_file']
        file_data = file.read()
        response = upload_document(pan_number, file_data)
        return jsonify(response=response)
    return render_template('document_scan.html')

@app.route('/call-log', methods=['GET', 'POST'])
def call_log():
    if request.method == 'POST':
        file = request.files['call_log_file']
        # Use a cross-platform temporary directory
        file_path = os.path.join(tempfile.gettempdir(), file.filename)
        file.save(file_path)
        response = upload_call_log(file_path)
        return jsonify(response=response)
    return render_template('call_log.html')

@app.route('/reports', methods=['GET', 'POST'])
def reports():
    if request.method == 'POST':
        report_type = request.form['report_type']
        response = generate_report(report_type)
        return jsonify(response=response)
    return render_template('reports.html')

def submit_expense(amount, description, date):
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        month = date_obj.strftime("%B")
        year = date_obj.year

        # Use user-specific directories for compatibility
        dir_path = os.path.join(os.path.expanduser("~"), "Documents", f"{month}_{year}")
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
        filename = os.path.join(dir_path, f"expenses_{date}.xlsx")

        if os.path.exists(filename):
            workbook = openpyxl.load_workbook(filename)
        else:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)

        if "Expenses" not in workbook.sheetnames:
            sheet = workbook.create_sheet(title='Expenses')
            sheet.append(['Amount', 'Description', 'Date'])
        else:
            sheet = workbook['Expenses']

        sheet.append([amount, description, date])
        workbook.save(filename)

        return "Expense saved successfully."
    except Exception as e:
        return f"An error occurred: {e}"

def generate_report(report_type):
    try:
        report_dir = os.path.join(os.path.expanduser("~"), "Documents", "Reports")
        if not os.path.exists(report_dir):
            os.makedirs(report_dir)

        filename = os.path.join(report_dir, f"{report_type}_report.txt")
        with open(filename, "w") as file:
            file.write(f"Report: {report_type.capitalize()} Report\n")
            file.write(f"Generated on: {datetime.now()}\n")
            file.write("This is a placeholder report.")

        return f"{report_type.capitalize()} report generated successfully at {filename}."
    except Exception as e:
        return f"An error occurred: {e}"

def upload_document(pan_number, file_data):
    try:
        doc_dir = os.path.join(os.path.expanduser("~"), "Documents", "PAN_Cards")
        if not os.path.exists(doc_dir):
            os.makedirs(doc_dir)

        destination_path = os.path.join(doc_dir, f"{pan_number}.pdf")
        with open(destination_path, "wb") as file:
            file.write(file_data)

        return f"Document saved successfully as {pan_number}.pdf."
    except Exception as e:
        return f"An error occurred: {e}"

def upload_call_log(file_path):
    try:
        call_log_dir = os.path.join(os.path.expanduser("~"), "Documents", "Call_Logs")
        if not os.path.exists(call_log_dir):
            os.makedirs(call_log_dir)

        destination_path = os.path.join(call_log_dir, os.path.basename(file_path))
        os.rename(file_path, destination_path)

        return f"Call log saved successfully as {os.path.basename(file_path)}."
    except Exception as e:
        return f"An error occurred: {e}"

if __name__ == '__main__':
    app.run(debug=True)
